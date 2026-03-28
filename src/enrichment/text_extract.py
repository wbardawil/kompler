"""Text extraction from various document formats.

Extracts readable text from PDFs, DOCX, XLSX, TXT, CSV, etc.
This is the first step before any AI processing.
"""

import io
import logging

logger = logging.getLogger(__name__)


def extract_text(file_bytes: bytes, filename: str, mime_type: str) -> str:
    """Extract text content from a document file.

    Returns plain text suitable for AI processing.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    try:
        if ext == "pdf" or mime_type == "application/pdf":
            return _extract_pdf(file_bytes)
        elif ext == "docx":
            return _extract_docx(file_bytes)
        elif ext == "xlsx":
            return _extract_xlsx(file_bytes)
        elif ext in ("txt", "csv", "md", "log"):
            return file_bytes.decode("utf-8", errors="replace")
        elif ext in ("png", "jpg", "jpeg", "tiff", "bmp"):
            return _extract_image_ocr(file_bytes)
        else:
            # Try as plain text
            return file_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        logger.warning("Text extraction failed for %s: %s", filename, e)
        return ""


def _extract_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF using PyPDF2 or pdfplumber."""
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            pages = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
    except ImportError:
        # Fallback to PyPDF2
        try:
            from PyPDF2 import PdfReader

            reader = PdfReader(io.BytesIO(file_bytes))
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
        except ImportError:
            logger.error("No PDF library available. Install pdfplumber or PyPDF2.")
            return ""


def _extract_docx(file_bytes: bytes) -> str:
    """Extract text from DOCX."""
    try:
        from docx import Document

        doc = Document(io.BytesIO(file_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        logger.error("python-docx not installed. Install with: pip install python-docx")
        return ""


def _extract_xlsx(file_bytes: bytes) -> str:
    """Extract text from XLSX spreadsheet."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            lines.append(f"=== Sheet: {sheet.title} ===")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        return "\n".join(lines)
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return ""


def _extract_image_ocr(file_bytes: bytes) -> str:
    """OCR for images. Phase 1: basic pytesseract. Phase 3: Claude Vision."""
    try:
        import pytesseract
        from PIL import Image

        image = Image.open(io.BytesIO(file_bytes))
        return pytesseract.image_to_string(image)
    except ImportError:
        logger.warning("OCR not available. Install pytesseract and Pillow.")
        return "[OCR not available — install pytesseract]"
